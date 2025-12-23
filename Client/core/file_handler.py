import os
import json
import openpyxl
from .logger import global_logger
from .config_manager import global_config

class FileHandler:
    """
    文件处理类，负责处理Excel和JSON文件
    """
    
    def __init__(self):
        """
        初始化文件处理类
        """
        # 从配置文件读取参数
        self.excel_file = global_config.get('File', 'excel_file')
        self.json_file = global_config.get('File', 'json_file')
        self.results_dir = global_config.get('File', 'results_dir')
        
        # 确保结果目录存在
        self._ensure_dir_exists(self.results_dir)
    
    def _ensure_dir_exists(self, directory):
        """
        确保目录存在，不存在则创建
        
        参数:
            directory: 目录路径
        """
        if directory and not os.path.exists(directory):
            try:
                os.makedirs(directory)
                global_logger.info(f"创建目录成功: {directory}")
            except Exception as e:
                global_logger.error(f"创建目录失败: {e}")
    
    def read_excel(self, file_path=None, sheet_name=None):
        """
        读取Excel文件
        
        参数:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认为活动工作表
            
        返回:
            tuple: (sheet, max_row, max_column) - 工作表对象、最大行数、最大列数
        """
        try:
            # 如果未提供文件路径，使用配置文件中的默认值
            if not file_path:
                file_path = self.excel_file
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                global_logger.error(f"Excel文件不存在: {file_path}")
                return (None, 0, 0)
            
            # 设置文件路径
            self.set_file_paths(excel_file=file_path)
            
            # 加载Excel文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # 获取工作表
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    global_logger.error(f"工作表不存在: {sheet_name}")
                    return (None, 0, 0)
            else:
                sheet = wb.active
            
            max_row = sheet.max_row
            max_column = sheet.max_column
            
            global_logger.info(f"Excel文件加载成功: {file_path}, 工作表: {sheet.title}, 行数: {max_row}, 列数: {max_column}")
            
            return (sheet, max_row, max_column)
            
        except Exception as e:
            global_logger.error(f"读取Excel文件失败: {e}")
            return (None, 0, 0)
    
    def get_cell_value(self, sheet, row, col):
        """
        获取指定单元格的值
        
        参数:
            sheet: 工作表对象
            row: 行号
            col: 列号
            
        返回:
            any: 单元格的值
        """
        try:
            if not sheet:
                global_logger.error("工作表对象为None")
                return None
            
            cell_value = sheet.cell(row=row, column=col).value
            global_logger.debug(f"获取单元格值: 行 {row}, 列 {col}, 值: {cell_value}")
            
            return cell_value
            
        except Exception as e:
            global_logger.error(f"获取单元格值失败: {e}")
            return None
    
    def save_json(self, data, file_path=None, append=True):
        """
        保存JSON数据
        
        参数:
            data: 要保存的数据
            file_path: JSON文件路径
            append: 是否追加模式，True为追加，False为覆盖
            
        返回:
            bool: 保存是否成功
        """
        try:
            # 如果未提供文件路径，使用配置文件中的默认值
            if not file_path:
                file_path = os.path.join(self.results_dir, self.json_file)
            
            # 确保目录存在
            file_dir = os.path.dirname(file_path)
            self._ensure_dir_exists(file_dir)
            
            # 转换数据为JSON字符串
            data_str = json.dumps(data, ensure_ascii=False, indent=2)
            
            # 写入文件
            mode = 'a' if append else 'w'
            with open(file_path, mode, encoding='utf-8') as f:
                if append:
                    f.write(data_str + '\n')
                else:
                    f.write(data_str)
            
            # 如果是追加模式，读取并验证最后一行数据
            if append:
                last_data = self.read_last_json(file_path)
                if last_data == data:
                    global_logger.debug(f"保存的数据已验证，与最后一行一致")
                else:
                    global_logger.warning(f"保存的数据与最后一行不一致")
            
            global_logger.info(f"JSON数据保存成功: {file_path}")
            return True
            
        except Exception as e:
            global_logger.error(f"保存JSON数据失败: {e}")
            return False
    
    def read_last_json(self, file_path=None):
        """
        读取JSON文件的最后一行数据
        
        参数:
            file_path: JSON文件路径
            
        返回:
            dict: 最后一行数据
        """
        try:
            # 如果未提供文件路径，使用配置文件中的默认值
            if not file_path:
                file_path = os.path.join(self.results_dir, self.json_file)
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                global_logger.error(f"JSON文件不存在: {file_path}")
                return {}
            
            # 读取文件的最后一行
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                if not lines:
                    return {}
                
                # 解析最后一行
                last_line = lines[-1].strip()
                if not last_line:
                    return {}
                
                data = json.loads(last_line)
                global_logger.info(f"读取JSON文件最后一行数据成功: {file_path}")
                return data
                
        except json.JSONDecodeError as e:
            global_logger.error(f"解析JSON数据失败: {e}")
            return {}
        except Exception as e:
            global_logger.error(f"读取JSON文件失败: {e}")
            return {}
    
    def read_all_json(self, file_path=None):
        """
        读取JSON文件的所有数据
        
        参数:
            file_path: JSON文件路径
            
        返回:
            list: 所有数据列表
        """
        try:
            # 如果未提供文件路径，使用配置文件中的默认值
            if not file_path:
                file_path = os.path.join(self.results_dir, self.json_file)
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                global_logger.error(f"JSON文件不存在: {file_path}")
                return []
            
            # 读取所有数据
            data_list = []
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        data = json.loads(line)
                        data_list.append(data)
            
            global_logger.info(f"读取JSON文件所有数据成功: {file_path}, 数据条数: {len(data_list)}")
            return data_list
            
        except json.JSONDecodeError as e:
            global_logger.error(f"解析JSON数据失败: {e}")
            return []
        except Exception as e:
            global_logger.error(f"读取JSON文件失败: {e}")
            return []
    
    def read_to_dict(self, data_string):
        """
        将字符串转换为字典
        
        参数:
            data_string: 要转换的字符串，格式为"accuracy difference target_weight time"
            
        返回:
            dict: 转换后的字典
        """
        try:
            # 分割字符串
            read_string = data_string.split(' ')
            # 转换为浮点数列表
            read_list = [float(x) for x in read_string]
            # 定义标题列表
            title_list = ['accuracy', 'difference', 'target_weight', 'time']
            # 转换为字典
            result_dict = dict(zip(title_list, read_list))
            
            global_logger.debug(f"字符串转换为字典成功: {result_dict}")
            return result_dict
            
        except ValueError as e:
            global_logger.error(f"字符串转换为字典失败: 数值转换错误, {e}")
            return {}
        except Exception as e:
            global_logger.error(f"字符串转换为字典失败: {e}")
            return {}
    
    def set_file_paths(self, excel_file=None, json_file=None, results_dir=None):
        """
        设置文件路径
        
        参数:
            excel_file: Excel文件路径
            json_file: JSON文件路径
            results_dir: 结果目录
        """
        if excel_file:
            self.excel_file = excel_file
            global_config.set('File', 'excel_file', excel_file)
        
        if json_file:
            self.json_file = json_file
            global_config.set('File', 'json_file', json_file)
        
        if results_dir:
            self.results_dir = results_dir
            self._ensure_dir_exists(results_dir)
            global_config.set('File', 'results_dir', results_dir)
        
        global_logger.info(f"文件路径已更新 - Excel: {self.excel_file}, JSON: {self.json_file}, 结果目录: {self.results_dir}")
